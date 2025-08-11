import streamlit as st
import pandas as pd
from io import BytesIO
import time
import re

# Parser 임포트
from danawa import DanawaParser
from guidecom import GuidecomParser

# openpyxl 스타일 임포트
from openpyxl.styles import Font, Alignment, PatternFill

# --- 페이지 설정 ---
st.set_page_config(
    page_title="가격 비교 사이트",
    page_icon="🛒",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# --- CSS 스타일 (공통) ---
st.markdown("""
<style>
    .main { padding: 0rem 1rem; }
    .stApp { background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); }
    .main-header { background: linear-gradient(135deg, #2c3e50, #3498db); padding: 2rem; border-radius: 15px; margin-bottom: 2rem; text-align: center; color: white; box-shadow: 0 10px 30px rgba(0,0,0,0.2); }
    .search-container { background: white; padding: 2rem; border-radius: 15px; box-shadow: 0 10px 30px rgba(0,0,0,0.1); margin-bottom: 2rem; }
    .manufacturer-grid { background: #f8f9fa; padding: 1.5rem; border-radius: 10px; border: 1px solid #dee2e6; }
    .results-container { background: white; padding: 2rem; border-radius: 15px; box-shadow: 0 10px 30px rgba(0,0,0,0.1); }
    .metric-container { background: linear-gradient(135deg, #27ae60, #2ecc71); padding: 0.8rem; border-radius: 8px; color: white; text-align: center; margin: 0.3rem; font-size: 0.9rem; }
    .metric-container h2 { font-size: 1.5rem; margin-bottom: 0.2rem; }
    .metric-container p { font-size: 0.8rem; margin-bottom: 0; }
    .stRadio > div { flex-direction: row; justify-content: center; }
    .stButton > button { background: linear-gradient(135deg, #3498db, #2980b9); color: white; border: none; border-radius: 8px; padding: 0.75rem 2rem; font-weight: 600; box-shadow: 0 4px 15px rgba(52, 152, 219, 0.3); transition: all 0.3s ease; }
    .stButton > button:hover { box-shadow: 0 6px 20px rgba(52, 152, 219, 0.4); transform: translateY(-2px); }
</style>
""", unsafe_allow_html=True)

# --- 사이트 선택 ---
site_choice = st.radio(
    "**검색할 사이트를 선택하세요**",
    ('다나와', '가이드컴'),
    horizontal=True,
    key='site_selector'
)

# ====================================================================
# 다나와(Danawa) 로직 (원본 코드와 100% 동일하게 유지)
# ====================================================================
if site_choice == '다나와':
    st.markdown('''<div class="main-header"><h1>🛒 다나와 가격비교</h1><p>최저가 제품을 쉽고 빠르게 찾아보세요</p></div>''', unsafe_allow_html=True)

    # --- 세션 상태 초기화 (원본과 동일) ---
    if 'parser' not in st.session_state:
        st.session_state.parser = DanawaParser()
    if 'search_options' not in st.session_state:
        st.session_state.search_options = []
    if 'search_results' not in st.session_state:
        st.session_state.search_results = []
    if 'show_manufacturers' not in st.session_state:
        st.session_state.show_manufacturers = False

    # --- 검색 UI (원본과 동일) ---
    with st.container():
        st.markdown('<div class="search-container">', unsafe_allow_html=True)
        st.markdown("### 🔍 제품 검색")
        col1, col2 = st.columns([3, 1])
        with col1:
            search_keyword = st.text_input("검색할 제품명을 입력하세요", placeholder="예: 9a1, RTX 4090...", key="danawa_search_input")
        with col2:
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("🔍 검색 옵션 로드", key="danawa_load_options"):
                if search_keyword.strip():
                    with st.spinner('검색 옵션을 가져오는 중...'):
                        try:
                            options = st.session_state.parser.get_search_options(search_keyword.strip())
                            manufacturer_options = [opt for opt in options if opt['category'] == '제조사']
                            st.session_state.search_options = manufacturer_options
                            st.session_state.show_manufacturers = True
                            if not manufacturer_options:
                                st.warning("⚠️ 제조사 옵션을 찾을 수 없습니다.")
                        except Exception as e:
                            st.error(f"❌ 옵션 조회 중 오류: {str(e)}")
                else:
                    st.error("❌ 검색어를 입력해주세요.")
        st.markdown('</div>', unsafe_allow_html=True)

    # --- 제조사 선택 및 검색 (원본과 동일) ---
    if st.session_state.show_manufacturers and st.session_state.search_options:
        st.markdown("<hr>", unsafe_allow_html=True)
        st.markdown("### 🏭 제조사 선택 (다중선택 가능)")
        st.markdown('<div class="manufacturer-grid">', unsafe_allow_html=True)
        manufacturers = st.session_state.search_options
        cols = st.columns(min(4, len(manufacturers)))
        selected_manufacturers = []
        for i, m in enumerate(manufacturers):
            with cols[i % len(cols)]:
                if st.checkbox(m['name'], key=f"danawa_m_{m['code']}"):
                    selected_manufacturers.append(m['code'])
        st.markdown('</div>', unsafe_allow_html=True)
        
        col1, col2, _ = st.columns([2, 2, 1])
        with col1:
            if st.button("🛒 제품 검색하기", key="danawa_search_products"):
                if search_keyword.strip():
                    with st.spinner('제품을 검색하는 중...'):
                        try:
                            all_products = []
                            if selected_manufacturers:
                                for code in selected_manufacturers:
                                    results = st.session_state.parser.search_all_categories(search_keyword.strip(), f"maker={code}")
                                    all_products.extend([p for prods in results.values() for p in prods])
                            else:
                                results = st.session_state.parser.search_all_categories(search_keyword.strip())
                                all_products.extend([p for prods in results.values() for p in prods])
                            
                            seen_names = set()
                            unique_products = [p for p in all_products if p.name not in seen_names and not seen_names.add(p.name)]
                            def extract_price(p_str):
                                nums = re.findall(r'[0-9,]+', p_str)
                                return int(nums[0].replace(',', '')) if nums else 999999999
                            unique_products.sort(key=lambda p: extract_price(p.price))
                            st.session_state.search_results = unique_products
                            if not unique_products:
                                st.warning("⚠️ 검색 결과가 없습니다.")
                        except Exception as e:
                            st.error(f"❌ 검색 중 오류: {str(e)}")
                else:
                    st.error("❌ 검색어를 입력해주세요.")
        with col2:
            if st.button("🔄 선택 초기화", key="danawa_clear"):
                st.session_state.search_results = []
                st.session_state.show_manufacturers = False
                st.rerun()

    # --- 결과 표시 (원본과 동일) ---
    if st.session_state.search_results:
        with st.container():
            st.markdown('<div class="results-container">', unsafe_allow_html=True)
            col1, col2, col3 = st.columns([1, 1, 1])
            with col1:
                st.markdown(f'''<div class="metric-container"><h2>{len(st.session_state.search_results)}</h2><p>총 제품 수</p></div>''', unsafe_allow_html=True)
            with col2:
                st.markdown(f'''<div class="metric-container"><h2>{search_keyword}</h2><p>검색어</p></div>''', unsafe_allow_html=True)
            with col3:
                def create_excel_file_danawa():
                    data = [{'No': i+1, '제품명': p.name, '가격': p.price, '세부사양': p.specifications} for i, p in enumerate(st.session_state.search_results)]
                    df = pd.DataFrame(data)
                    output = BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        df.to_excel(writer, sheet_name='다나와 검색결과', index=False)
                        ws = writer.sheets['다나와 검색결과']
                        header_font = Font(bold=True, color="FFFFFF")
                        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                        header_alignment = Alignment(horizontal='center', vertical='center')
                        for col_num, cell in enumerate(ws[1], 1):
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = header_alignment
                        column_widths = [8, 50, 15, 80]
                        for i, width in enumerate(column_widths, 1):
                            ws.column_dimensions[ws.cell(1, i).column_letter].width = width
                    output.seek(0)
                    return output.getvalue()
                if st.button("📥 Excel 다운로드", key="danawa_download_excel"):
                    excel_data = create_excel_file_danawa()
                    st.download_button("💾 Excel 파일 다운로드", excel_data, f"{search_keyword}_다나와_결과.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="danawa_download_button")
            st.markdown("<hr>", unsafe_allow_html=True)
            st.markdown("### 📋 검색 결과")
            df_display = pd.DataFrame([{'No': i+1, '제품명': p.name, '가격': p.price, '세부사양': p.specifications} for i, p in enumerate(st.session_state.search_results)])
            st.dataframe(df_display, use_container_width=True, hide_index=True, column_config={"No": st.column_config.NumberColumn("No", width=60), "제품명": st.column_config.TextColumn("제품명", width=350), "가격": st.column_config.TextColumn("가격", width=120), "세부사양": st.column_config.TextColumn("세부사양")})
            st.markdown('</div>', unsafe_allow_html=True)

# ====================================================================
# 가이드컴(Guidecom) 로직
# ====================================================================
else:
    st.markdown('''<div class="main-header"><h1>🛒 가이드컴 가격비교</h1><p>최저가 제품을 쉽고 빠르게 찾아보세요</p></div>''', unsafe_allow_html=True)

    # --- 세션 상태 초기화 (가이드컴 전용) ---
    if 'guidecom_parser' not in st.session_state:
        st.session_state.guidecom_parser = GuidecomParser()
    if 'guidecom_search_results' not in st.session_state:
        st.session_state.guidecom_search_results = []

    # --- 검색 UI ---
    with st.container():
        st.markdown('<div class="search-container">', unsafe_allow_html=True)
        st.markdown("### 🔍 제품 검색")
        search_keyword_guidecom = st.text_input("검색할 제품명을 입력하세요", placeholder="예: 라이젠 7800X3D...", key="guidecom_search_input")
        if st.button("🛒 제품 검색하기", key="guidecom_search_products"):
            if search_keyword_guidecom.strip():
                with st.spinner('제품을 검색하는 중...'):
                    results = st.session_state.guidecom_parser.search_all_categories(search_keyword_guidecom.strip())
                    all_products = [p for prods in results.values() for p in prods]
                    seen_names = set()
                    unique_products = [p for p in all_products if p.name not in seen_names and not seen_names.add(p.name)]
                    def extract_price(p_str):
                        nums = re.findall(r'[0-9,]+', p_str)
                        return int(nums[0].replace(',', '')) if nums else 999999999
                    unique_products.sort(key=lambda p: extract_price(p.price))
                    st.session_state.guidecom_search_results = unique_products
                    if not unique_products:
                        st.warning("⚠️ 검색 결과가 없습니다.")
            else:
                st.error("❌ 검색어를 입력해주세요.")
        st.markdown('</div>', unsafe_allow_html=True)

    # --- 결과 표시 UI ---
    if st.session_state.guidecom_search_results:
        with st.container():
            st.markdown('<div class="results-container">', unsafe_allow_html=True)
            col1, col2, col3 = st.columns([1, 1, 1])
            with col1:
                st.markdown(f'''<div class="metric-container"><h2>{len(st.session_state.guidecom_search_results)}</h2><p>총 제품 수</p></div>''', unsafe_allow_html=True)
            with col2:
                st.markdown(f'''<div class="metric-container"><h2>{search_keyword_guidecom}</h2><p>검색어</p></div>''', unsafe_allow_html=True)
            with col3:
                def create_excel_file_guidecom():
                    data = [{'No': i+1, '제품명': p.name, '가격': p.price, '세부사양': p.specifications} for i, p in enumerate(st.session_state.guidecom_search_results)]
                    df = pd.DataFrame(data)
                    output = BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        df.to_excel(writer, sheet_name='가이드컴 검색결과', index=False)
                        ws = writer.sheets['가이드컴 검색결과']
                        header_font = Font(bold=True, color="FFFFFF")
                        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                        header_alignment = Alignment(horizontal='center', vertical='center')
                        for col_num, cell in enumerate(ws[1], 1):
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = header_alignment
                        column_widths = [8, 50, 15, 80]
                        for i, width in enumerate(column_widths, 1):
                            ws.column_dimensions[ws.cell(1, i).column_letter].width = width
                    output.seek(0)
                    return output.getvalue()
                if st.button("📥 Excel 다운로드", key="guidecom_download_excel"):
                    excel_data = create_excel_file_guidecom()
                    st.download_button("💾 Excel 파일 다운로드", excel_data, f"{search_keyword_guidecom}_가이드컴_결과.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="guidecom_download_button")
            st.markdown("<hr>", unsafe_allow_html=True)
            st.markdown("### 📋 검색 결과")
            df_display = pd.DataFrame([{'No': i+1, '제품명': p.name, '가격': p.price, '세부사양': p.specifications} for i, p in enumerate(st.session_state.guidecom_search_results)])
            st.dataframe(df_display, use_container_width=True, hide_index=True, column_config={"No": st.column_config.NumberColumn("No", width=60), "제품명": st.column_config.TextColumn("제품명", width=350), "가격": st.column_config.TextColumn("가격", width=120), "세부사양": st.column_config.TextColumn("세부사양")})
            st.markdown('</div>', unsafe_allow_html=True)

# --- 공통 푸터 ---
st.markdown("---")
st.markdown(f'''<div style='text-align: center; color: white; padding: 1rem;'><p>🛒 {site_choice} 가격비교 웹앱 | Made with ❤️ using Streamlit</p></div>''', unsafe_allow_html=True)
