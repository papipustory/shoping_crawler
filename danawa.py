import requests
from bs4 import BeautifulSoup
import time
import json
import urllib.parse
from dataclasses import dataclass
from typing import List, Dict, Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

@dataclass
class Product:
    name: str
    price: str
    specifications: str

class DanawaParser:
    def __init__(self):
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (iPhone; CPU iPhone OS 14_7_1 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.1.2 Mobile/15E148 Safari/604.1'
        })
        self.base_url = "https://search.danawa.com/mobile/dsearch.php"
    
    def get_search_options(self, keyword: str) -> List[Dict[str, str]]:
        """
        검색 페이지에서 옵션명들을 가져옵니다.
        
        Args:
            keyword: 검색할 제품명
            
        Returns:
            옵션 리스트 [{'name': '옵션명', 'code': '옵션코드'}, ...]
        """
        params = {
            'keyword': keyword
        }
        
        try:
            response = self.session.get(self.base_url, params=params)
            response.raise_for_status()
            
            soup = BeautifulSoup(response.text, 'html.parser')
            options = []
            
            # 제조사/브랜드 탭에서 제조사 찾기 (검색된 키워드에 맞는 제조사들)
            maker_tab = soup.find('div', id='makerBrandTab')
            if maker_tab:
                # 제조사 버튼들 찾기 - 정확한 클래스명 사용 
                buttons = maker_tab.find_all('button', class_='button__option')
                
                print(f"찾은 제조사 버튼 수: {len(buttons)}")
                
                # makerBrandTab에서 제조사 이름들 수집
                brand_names = []
                for button in buttons:
                    option_name = button.get('data-optionname')
                    if option_name:
                        brand_names.append(option_name)
                        print(f"makerBrandTab에서 찾은 제조사: {option_name}")
                
                # makerOptionTab에서 실제 코드 찾기
                option_tab = soup.find('div', id='makerOptionTab')
                if option_tab:
                    option_buttons = option_tab.find_all('button', class_='button__option')
                    print(f"makerOptionTab에서 찾은 버튼 수: {len(option_buttons)}")
                    
                    for opt_button in option_buttons:
                        opt_name = opt_button.get('data-optionname')
                        opt_code = opt_button.get('data-optioncode')
                        
                        print(f"makerOptionTab - 제조사: {opt_name}, 코드: {opt_code}")
                        
                        # makerBrandTab에 있는 제조사만 추가
                        if opt_name in brand_names and opt_code and opt_code.isdigit():
                            options.append({
                                'category': '제조사',
                                'name': opt_name,
                                'code': opt_code
                            })
                            print(f"매핑 완료: {opt_name} -> {opt_code}")
                else:
                    print("makerOptionTab을 찾을 수 없음")
            
            # 옵션을 찾지 못한 경우 기본 제조사 목록 사용
            if not options:
                default_manufacturers = [
                    {'category': '제조사', 'name': '삼성전자', 'code': '185'},
                    {'category': '제조사', 'name': 'LG전자', 'code': '21'},
                    {'category': '제조사', 'name': 'ASUS', 'code': '17'},
                    {'category': '제조사', 'name': 'MSI', 'code': '143'},
                    {'category': '제조사', 'name': 'GIGABYTE', 'code': '399'},
                    {'category': '제조사', 'name': 'EVGA', 'code': '3148'},
                    {'category': '제조사', 'name': '조텍', 'code': '3142'},
                    {'category': '제조사', 'name': '갤럭시', 'code': '3154'},
                    {'category': '제조사', 'name': '인텔', 'code': '16'},
                    {'category': '제조사', 'name': 'AMD', 'code': '238'},
                    {'category': '제조사', 'name': '웨스턴디지털', 'code': '22'},
                    {'category': '제조사', 'name': '시게이트', 'code': '24'},
                ]
                return default_manufacturers
            
            return options[:20]  # 최대 20개로 제한
            
        except Exception as e:
            print(f"옵션 조회 중 오류 발생: {e}")
            # 오류 발생시 기본 제조사 목록 반환
            return [
                {'category': '제조사', 'name': '삼성전자', 'code': '185'},
                {'category': '제조사', 'name': 'LG전자', 'code': '21'},
                {'category': '제조사', 'name': 'ASUS', 'code': '17'},
                {'category': '제조사', 'name': 'MSI', 'code': '143'},
                {'category': '제조사', 'name': 'GIGABYTE', 'code': '399'},
                {'category': '제조사', 'name': '웨스턴디지털', 'code': '22'},
                {'category': '제조사', 'name': '시게이트', 'code': '24'},
                {'category': '제조사', 'name': '인텔', 'code': '16'},
                {'category': '제조사', 'name': 'AMD', 'code': '238'},
            ]
    
    def search_products(self, keyword: str, sort_type: str = "saveDESC", limit: int = 5, option_filter: str = None) -> List[Product]:
        """
        다나와에서 제품을 검색합니다.
        
        Args:
            keyword: 검색할 제품명
            sort_type: 정렬 방식 
                - "priceASC": 낮은가격순
                - "saveDESC": 인기상품순
                - "opinionDESC": 상품평많은순
            limit: 가져올 제품 개수
            option_filter: 옵션 필터 (예: "maker=3148")
        """
        params = {
            'keyword': keyword,
            'sort': sort_type
        }
        
        # 옵션 필터가 있으면 추가 (예: maker=3148)
        if option_filter:
            key, value = option_filter.split('=')
            
            # 다나와 모바일에서 사용하는 제조사 필터 파라미터
            if key == 'maker':
                # maker 파라미터 사용 (예: maker=702 또는 maker=4213%2C702)
                params['maker'] = value
                print(f"제조사 필터: maker={value} 적용")
                
            params[key] = value
            print(f"제조사 필터 적용: {key}={value}")
            print(f"추가 제조사 파라미터들도 적용")
        
        print(f"검색 URL: {self.base_url}")
        print(f"검색 파라미터: {params}")
        
        try:
            response = self.session.get(self.base_url, params=params)
            response.raise_for_status()
            print(f"응답 상태: {response.status_code}")
            print(f"최종 URL: {response.url}")
            
            soup = BeautifulSoup(response.text, 'html.parser')
            products = []
            
            # 제품 목록 파싱 - 모바일 버전 (올바른 구조)
            product_items = soup.find_all('li', class_='goods-list__item')
            
            # 제품 아이템이 없는 경우 다른 클래스들 시도
            if not product_items:
                alternative_classes = ['product-item', 'goods-item', 'list-item', 'item']
                for cls in alternative_classes:
                    product_items = soup.find_all('li', class_=cls)
                    if product_items:
                        break
                
                # li 태그가 아닌 div 태그로 된 경우도 시도
                if not product_items:
                    for cls in ['goods-list__item', 'product-item', 'goods-item']:
                        product_items = soup.find_all('div', class_=cls)
                        if product_items:
                            break
            
            print(f"찾은 제품 아이템 수: {len(product_items)}")
            
            parsed_count = 0
            for i, item in enumerate(product_items[:limit * 2]):  # 여유있게 더 많이 시도
                try:
                    print(f"제품 {i+1} 파싱 시도...")
                    product = self._parse_product_item(item)
                    if product:
                        print(f"제품 파싱 성공: {product.name[:50]}...")
                        products.append(product)
                        parsed_count += 1
                        if parsed_count >= limit:
                            break
                    else:
                        print(f"제품 {i+1} 파싱 실패 - 유효하지 않은 데이터")
                    time.sleep(0.3)  # 요청 간격 조절 (조금 빠르게)
                except Exception as e:
                    print(f"개별 제품 파싱 중 오류: {e}")
                    continue
            
            return products
            
        except Exception as e:
            print(f"검색 중 오류 발생: {e}")
            return []
    
    def _parse_product_item(self, item) -> Optional[Product]:
        """제품 아이템에서 정보를 추출합니다."""
        try:
            # 제품명 추출 - 여러 방법 시도
            name = "정보 없음"
            
            # 방법 1: goods-list__title 클래스
            name_elem = item.find('span', class_='goods-list__title')
            if name_elem:
                # <br> 태그를 공백으로 변환
                for br in name_elem.find_all('br'):
                    br.replace_with(' ')
                name = name_elem.get_text(strip=True)
                # 연속된 공백을 하나로 압축
                name = ' '.join(name.split())
                print(f"제품명 찾음: '{name}'")
            else:
                print("goods-list__title 클래스를 찾을 수 없음")
                # 방법 2: a 태그의 title 속성
                link_elem = item.find('a', title=True)
                if link_elem:
                    name = link_elem.get('title', '').strip()
                else:
                    # 방법 3: 일반적인 제품명 클래스들 시도
                    for cls in ['title', 'product-title', 'goods-title', 'item-title']:
                        elem = item.find(class_=cls)
                        if elem and elem.text.strip():
                            name = elem.text.strip()
                            break
            
            # 가격 추출 - 여러 방법 시도
            price = "가격 문의"
            
            # 방법 1: goods-list__price > em.num (실제 구조)
            price_div = item.find('div', class_='goods-list__price')
            if price_div:
                # 'num' 클래스 시도
                number_elem = price_div.find('em', class_='num')
                if not number_elem:
                    # 'number' 클래스도 시도 (하위 호환)
                    number_elem = price_div.find('em', class_='number')
                
                if number_elem and number_elem.text.strip():
                    price_text = number_elem.text.strip()
                    if price_text and price_text != '-':
                        price = price_text + "원"
                        print(f"가격 찾음: '{price}'")
            
            # 방법 2: 다른 가격 클래스들 시도
            if price == "가격 문의":
                for cls in ['price', 'cost', 'money', 'won']:
                    price_elem = item.find(class_=cls)
                    if price_elem and price_elem.text.strip():
                        price_text = price_elem.text.strip()
                        if price_text and price_text != '-' and '원' not in price_text:
                            price = price_text + "원"
                            break
                        elif '원' in price_text:
                            price = price_text
                            break
            
            # 세부 사양 추출
            specifications = self._get_specifications_from_item(item)
            
            # 유효성 검사
            if not name or name.strip() == "정보 없음" or len(name.strip()) < 2:
                return None
            
            return Product(
                name=name.strip(),
                price=price,
                specifications=specifications
            )
            
        except Exception as e:
            print(f"제품 파싱 중 오류: {e}")
            return None
    
    def _get_specifications_from_item(self, item) -> str:
        """검색 결과 아이템에서 직접 사양 정보를 추출합니다."""
        try:
            specs = []
            
            # 방법 1: 해당 li 요소나 인근에서 spec-box__inner 찾기
            # li 요소 자체에서 찾기
            spec_inner = item.find('div', class_='spec-box__inner')
            
            # li 요소에 없으면 부모나 형제 요소에서 찾기
            if not spec_inner:
                parent = item.parent
                if parent:
                    # 부모의 모든 자식에서 찾기
                    spec_inner = parent.find('div', class_='spec-box__inner')
                
                # 여전히 없으면 더 넓은 범위에서 찾기
                if not spec_inner:
                    # productItem ID를 기반으로 연관 사양 찾기
                    item_id = item.get('id')
                    if item_id:
                        # 같은 페이지에서 관련 사양 정보 찾기
                        soup = item.find_parent().find_parent() if item.parent else None
                        if soup:
                            spec_inner = soup.find('div', class_='spec-box__inner')
            
            if spec_inner:
                spec_spans = spec_inner.find_all('span')
                for span in spec_spans:
                    # slash 클래스가 아닌 경우만 추출
                    span_classes = span.get('class', [])
                    if 'slash' not in span_classes:
                        text = span.text.strip()
                        if text and len(text) > 1 and text not in ['/', '|', '-']:
                            specs.append(text)
                            
                print(f"사양 정보 찾음: {len(specs)}개")
            
            # 방법 2: 다른 사양 관련 클래스들 시도
            if not specs:
                spec_classes = ['spec', 'specification', 'details', 'info', 'feature']
                for cls in spec_classes:
                    spec_elem = item.find(class_=cls)
                    if spec_elem:
                        # 내부 텍스트들 수집
                        text_elements = spec_elem.find_all(text=True)
                        for text in text_elements:
                            clean_text = text.strip()
                            if clean_text and len(clean_text) > 1 and clean_text not in ['/', '|', '-']:
                                specs.append(clean_text)
                        if specs:
                            break
            
            # 방법 3: 일반적인 정보 추출 (링크나 버튼의 title 속성 등)
            if not specs:
                for elem in item.find_all(['a', 'button', 'div'], title=True):
                    title_text = elem.get('title', '').strip()
                    if title_text and len(title_text) > 5:  # 충분히 긴 텍스트만
                        specs.append(title_text)
                        break
            
            # 중복 제거 및 필터링
            if specs:
                # 중복 제거
                unique_specs = []
                seen = set()
                for spec in specs[:15]:  # 최대 15개
                    if spec.lower() not in seen and len(spec.strip()) > 1:
                        unique_specs.append(spec.strip())
                        seen.add(spec.lower())
                
                if unique_specs:
                    return " / ".join(unique_specs[:10])
            
            return "사양 정보 없음"
            
        except Exception as e:
            print(f"사양 정보 추출 실패: {e}")
            return "사양 정보 없음"

    def _get_product_specifications(self, product_url: str) -> str:
        """레거시 코드 - 사용하지 않음"""
        return "사양 정보 없음"
    
    def search_all_categories(self, keyword: str, option_filter: str = None) -> Dict[str, List[Product]]:
        """모든 카테고리별로 제품을 검색합니다."""
        # 인기상품 5개, 상품평많은순 5개
        popular_products = self.search_products(keyword, "saveDESC", 5, option_filter)
        review_products = self.search_products(keyword, "opinionDESC", 5, option_filter)
        
        # 각 카테고리별로 중복 제거 (각 카테고리에 최소 3개씩 보장)
        def remove_duplicates_keep_min(products, min_count=3):
            seen_names = set()
            unique = []
            for product in products:
                if product.name not in seen_names or len(unique) < min_count:
                    seen_names.add(product.name)
                    unique.append(product)
                if len(unique) >= 5:  # 최대 5개
                    break
            return unique
        
        # 인기상품순 중복 제거
        popular_unique = remove_duplicates_keep_min(popular_products)
        popular_names = {p.name for p in popular_unique}
        
        # 상품평많은순에서 인기상품과 중복되지 않는 것만 선택
        review_unique = []
        for product in review_products:
            if product.name not in popular_names:
                review_unique.append(product)
            if len(review_unique) >= 5:
                break
        
        # 상품평많은순이 적으면 인기상품에서 채우기
        if len(review_unique) < 3:
            for product in review_products:
                if product.name in popular_names and len(review_unique) < 5:
                    review_unique.append(product)
        
        results = {
            "인기상품순": popular_unique,
            "상품평많은순": review_unique
        }
        
        return results
    
    def print_results(self, results: Dict[str, List[Product]]):
        """검색 결과를 출력합니다."""
        for category, products in results.items():
            print(f"\n{'='*50}")
            print(f"{category} ({len(products)}개)")
            print('='*50)
            
            for i, product in enumerate(products, 1):
                print(f"\n{i}. {product.name}")
                print(f"   가격: {product.price}")
                print(f"   세부 사양: {product.specifications}")
    
    def save_to_excel(self, results: Dict[str, List[Product]], keyword: str):
        """검색 결과를 엑셀 파일로 저장합니다."""
        try:
            # 모든 제품 데이터 수집
            all_products = []
            for category, products in results.items():
                for product in products:
                    all_products.append(product)
            
            # 가격순 정렬 (낮은가격부터 오름차순)
            def extract_price_number(price_str):
                import re
                # "원" 제거하고 콤마 제거하여 숫자만 추출
                numbers = re.findall(r'[0-9,]+', price_str)
                if numbers:
                    return int(numbers[0].replace(',', ''))
                return 999999999  # 가격 정보 없으면 맨 뒤로
            
            all_products.sort(key=lambda p: extract_price_number(p.price))
            
            # 데이터 준비
            data = []
            for i, product in enumerate(all_products, 1):
                data.append({
                    'No': i,
                    '제품명': product.name,
                    '가격': product.price,
                    '세부사양': product.specifications
                })
            
            # DataFrame 생성
            df = pd.DataFrame(data)
            
            # 엑셀 파일 생성
            filename = f"{keyword}_danawa_results.xlsx"
            
            # openpyxl을 사용하여 세밀한 조정
            wb = Workbook()
            ws = wb.active
            ws.title = "다나와 검색결과"
            
            # 헤더 작성
            headers = ['No', '제품명', '가격', '세부사양']
            ws.append(headers)
            
            # 헤더 스타일 설정
            header_font = Font(bold=True)
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = header_font
                cell.alignment = header_alignment
            
            # 데이터 추가
            for row_data in data:
                ws.append([row_data['No'], row_data['제품명'], row_data['가격'], row_data['세부사양']])
            
            # 셀 정렬 설정
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    if cell.column == 1:  # No 열
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    elif cell.column == 4:  # 세부사양 열
                        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # 칼럼 너비 자동 조정
            self._adjust_column_widths(ws)
            
            # 파일 저장
            wb.save(filename)
            print(f"엑셀 파일이 '{filename}'에 저장되었습니다.")
            
        except Exception as e:
            print(f"엑셀 파일 저장 중 오류 발생: {e}")
    
    def _adjust_column_widths(self, worksheet):
        """칼럼 너비를 내용에 맞게 자동 조정합니다."""
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        # 세부사양 칼럼의 경우 줄바꿈 고려
                        if cell.column == 4 and '\n' in str(cell.value):
                            lines = str(cell.value).split('\n')
                            cell_length = max(len(line) for line in lines)
                        else:
                            cell_length = len(str(cell.value))
                        
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            # 칼럼별 최대/최소 너비 설정
            if column_letter == 'A':  # No 칼럼
                adjusted_width = max(5, min(max_length + 2, 8))
            elif column_letter == 'B':  # 제품명 칼럼
                adjusted_width = max(20, min(max_length + 2, 60))
            elif column_letter == 'C':  # 가격 칼럼
                adjusted_width = max(12, min(max_length + 2, 20))
            elif column_letter == 'D':  # 세부사양 칼럼
                adjusted_width = max(30, min(max_length + 2, 80))
            else:
                adjusted_width = max_length + 2
            
            worksheet.column_dimensions[column_letter].width = adjusted_width

def main():
    parser = DanawaParser()
    
    # 사용자 입력
    keyword = input("검색할 제품명을 입력하세요: ").strip()
    
    if not keyword:
        print("제품명을 입력해주세요.")
        return
    
    print(f"\n'{keyword}' 검색 옵션을 가져오는 중...")
    
    # 검색 옵션 가져오기
    options = parser.get_search_options(keyword)
    
    option_filter = None
    if options:
        print("\n=== 검색 옵션 ===")
        print("0. 옵션 필터 없이 검색")
        
        # 카테고리별로 그룹화
        categories = {}
        for i, option in enumerate(options, 1):
            category = option['category']
            if category not in categories:
                categories[category] = []
            categories[category].append((i, option))
        
        # 옵션 출력
        for category, items in categories.items():
            if category in ["제조사", "제조사/브랜드"]:  # 제조사만 표시
                print(f"\n[{category}]")
                for idx, option in items:
                    print(f"{idx}. {option['name']}")
        
        # 사용자 선택
        try:
            choice = int(input("\n원하는 옵션 번호를 선택하세요 (0: 필터 없음): "))
            if choice > 0 and choice <= len(options):
                selected_option = options[choice - 1]
                option_filter = f"maker={selected_option['code']}"
                print(f"\n선택된 옵션: {selected_option['name']}")
            elif choice == 0:
                print("\n옵션 필터 없이 검색합니다.")
            else:
                print("\n잘못된 선택입니다. 필터 없이 검색합니다.")
        except ValueError:
            print("\n잘못된 입력입니다. 필터 없이 검색합니다.")
    else:
        print("\n검색 옵션을 찾을 수 없습니다.")
    
    print(f"\n'{keyword}' 검색 중...")
    
    # 모든 카테고리 검색
    results = parser.search_all_categories(keyword, option_filter)
    
    # 결과 출력
    parser.print_results(results)
    
    # 엑셀로 자동 저장 (최적화)
    parser.save_to_excel(results, keyword)

if __name__ == "__main__":
    main()
