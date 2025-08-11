import requests
from bs4 import BeautifulSoup
import time
from dataclasses import dataclass
from typing import List, Dict, Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import re

@dataclass
class Product:
    name: str
    price: str
    specifications: str

class GuidecomParser:
    def __init__(self):
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36'
        })
        self.base_url = "https://www.guidecom.co.kr/shop/search.php"

    def get_search_options(self, keyword: str) -> List[Dict[str, str]]:
        """
        가이드컴은 다나와처럼 동적 검색 옵션을 제공하지 않는 것으로 보입니다.
        이 기능이 필요하다면, 어떤 옵션을 가져와야 하는지 더 자세한 정보가 필요합니다.
        """
        print("가이드컴에서는 이 기능을 지원하지 않거나, 다른 방식의 구현이 필요합니다.")
        return []

    def search_products(self, keyword: str, sort_type: str = "hit", limit: int = 5) -> List[Product]:
        """
        가이드컴에서 제품을 검색합니다.
        
        Args:
            keyword: 검색할 제품명
            sort_type: 정렬 방식 
                - "hit": 인기상품순
                - "new": 신상품순
                - "pricea": 낮은가격순
                - "priced": 높은가격순
                - "name": 상품명순
            limit: 가져올 제품 개수
        """
        params = {
            'q': keyword,
            'sort': sort_type
        }
        
        try:
            response = self.session.get(self.base_url, params=params)
            response.raise_for_status()
            
            soup = BeautifulSoup(response.text, 'html.parser')
            products = []
            
            product_list_ul = soup.find('ul', class_='goods-list')
            if not product_list_ul:
                return []
                
            product_items = product_list_ul.find_all('li')
            
            for item in product_items[:limit]:
                product = self._parse_product_item(item)
                if product:
                    products.append(product)
                    time.sleep(0.2)
            
            return products
            
        except Exception as e:
            print(f"검색 중 오류 발생: {e}")
            return []

    def _parse_product_item(self, item) -> Optional[Product]:
        """제품 아이템에서 정보를 추출합니다."""
        try:
            name_elem = item.select_one('.goods-list-name strong')
            name = name_elem.text.strip() if name_elem else "정보 없음"
            
            price_elem = item.select_one('.goods-list-price .cost')
            price = price_elem.text.strip() + "원" if price_elem and price_elem.text.strip() else "가격 문의"
            
            spec_elem = item.select_one('.goods-list-spec')
            specifications = spec_elem.text.strip().replace('\n', ' / ') if spec_elem else "사양 정보 없음"
            
            return Product(
                name=name,
                price=price,
                specifications=specifications
            )
            
        except Exception as e:
            print(f"제품 파싱 중 오류: {e}")
            return None

    def search_all_categories(self, keyword: str) -> Dict[str, List[Product]]:
        """주요 카테고리별로 제품을 검색합니다."""
        popular_products = self.search_products(keyword, "hit", 5)
        new_products = self.search_products(keyword, "new", 5)
        
        def remove_duplicates(products, max_count=5):
            seen_names = set()
            unique = []
            for product in products:
                if product.name not in seen_names:
                    seen_names.add(product.name)
                    unique.append(product)
                if len(unique) >= max_count:
                    break
            return unique

        popular_unique = remove_duplicates(popular_products)
        popular_names = {p.name for p in popular_unique}

        new_unique = []
        for product in new_products:
            if product.name not in popular_names:
                new_unique.append(product)
            if len(new_unique) >= 5:
                break
        
        results = {
            "인기상품순": popular_unique,
            "신상품순": new_unique
        }
        
        return results

    def print_results(self, results: Dict[str, List[Product]]):
        """검색 결과를 출력합니다."""
        for category, products in results.items():
            print(f"\n{'='*50}")
            print(f"{category} ({len(products)}개)")
            print('='*50)
            
            for i, product in enumerate(products, 1):
                print(f"\n{i}. {product.name}")
                print(f"   가격: {product.price}")
                print(f"   세부 사양: {product.specifications}")

    def _adjust_column_widths(self, worksheet):
        """칼럼 너비를 내용에 맞게 자동 조정합니다."""
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        if cell.column == 4 and '\n' in str(cell.value):
                            lines = str(cell.value).split('\n')
                            cell_length = max(len(line) for line in lines)
                        else:
                            cell_length = len(str(cell.value))
                        
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            if column_letter == 'A':
                adjusted_width = max(5, min(max_length + 2, 8))
            elif column_letter == 'B':
                adjusted_width = max(20, min(max_length + 2, 60))
            elif column_letter == 'C':
                adjusted_width = max(12, min(max_length + 2, 20))
            elif column_letter == 'D':
                adjusted_width = max(30, min(max_length + 2, 80))
            else:
                adjusted_width = max_length + 2
            
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def _prepare_excel_data(self, results: Dict[str, List[Product]]):
        """엑셀 저장을 위한 데이터를 준비하고 중복을 제거합니다."""
        all_products = []
        for products in results.values():
            all_products.extend(products)
        
        def extract_price_number(price_str):
            numbers = re.findall(r'[0-9,]+', price_str)
            if numbers:
                return int(numbers[0].replace(',', ''))
            return 999999999
        
        all_products.sort(key=lambda p: extract_price_number(p.price))
        
        data = []
        seen_names = set()
        for product in all_products:
            if product.name not in seen_names:
                data.append({
                    '제품명': product.name,
                    '가격': product.price,
                    '세부사양': product.specifications
                })
                seen_names.add(product.name)

        for i, row in enumerate(data, 1):
            row['No'] = i
            
        return pd.DataFrame(data, columns=['No', '제품명', '가격', '세부사양'])

    def save_to_excel(self, results: Dict[str, List[Product]], keyword: str):
        """검색 결과를 엑셀 파일로 저장합니다."""
        try:
            df = self._prepare_excel_data(results)
            if df.empty:
                print("저장할 데이터가 없습니다.")
                return

            filename = f"{keyword}_guidecom_results.xlsx"
            
            wb = Workbook()
            ws = wb.active
            ws.title = "가이드컴 검색결과"
            
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)

            header_font = Font(bold=True)
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            for col_num, header in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = header_font
                cell.alignment = header_alignment
            
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    alignment_kwargs = {'vertical': 'center'}
                    if cell.column == 1:
                        alignment_kwargs['horizontal'] = 'center'
                    elif cell.column == 4:
                        alignment_kwargs['horizontal'] = 'left'
                        alignment_kwargs['vertical'] = 'top'
                        alignment_kwargs['wrap_text'] = True
                    else:
                        alignment_kwargs['horizontal'] = 'left'
                    cell.alignment = Alignment(**alignment_kwargs)
            
            self._adjust_column_widths(ws)
            
            wb.save(filename)
            print(f"엑셀 파일이 '{filename}'에 저장되었습니다.")
            
        except Exception as e:
            print(f"엑셀 파일 저장 중 오류 발생: {e}")

    def save_to_excel_web(self, results: Dict[str, List[Product]], keyword: str, filepath: str):
        """웹 버전용 엑셀 파일 저장"""
        try:
            df = self._prepare_excel_data(results)
            if df.empty:
                raise Exception("저장할 데이터가 없습니다.")

            wb = Workbook()
            ws = wb.active
            ws.title = "가이드컴 검색결과"
            
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)

            header_font = Font(bold=True)
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            for col_num, header in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = header_font
                cell.alignment = header_alignment
            
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    alignment_kwargs = {'vertical': 'center'}
                    if cell.column == 1:
                        alignment_kwargs['horizontal'] = 'center'
                    elif cell.column == 4:
                        alignment_kwargs['horizontal'] = 'left'
                        alignment_kwargs['vertical'] = 'top'
                        alignment_kwargs['wrap_text'] = True
                    else:
                        alignment_kwargs['horizontal'] = 'left'
                    cell.alignment = Alignment(**alignment_kwargs)
            
            self._adjust_column_widths(ws)
            
            wb.save(filepath)
            
        except Exception as e:
            raise Exception(f"엑셀 파일 저장 중 오류 발생: {e}")

def main():
    parser = GuidecomParser()
    
    keyword = input("검색할 제품명을 입력하세요: ").strip()
    
    if not keyword:
        print("제품명을 입력해주세요.")
        return
    
    print(f"\n'{keyword}' 검색 중...")
    
    results = parser.search_all_categories(keyword)
    
    if not any(results.values()):
        print("\n검색 결과가 없습니다.")
        return

    parser.print_results(results)
    
    parser.save_to_excel(results, keyword)

if __name__ == "__main__":
    main()
